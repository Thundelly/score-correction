from PyQt5 import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from calender_dialog import *
import datetime
import pathlib
import os
from shutil import *
from openpyxl import *
from openpyxl.drawing.image import Image
from openpyxl import styles
import openpyxl
import copy
import platform
import subprocess

sat_report_path = os.path.join("template_reports", "sat_2018_analysis_report.xlsx")
psat_report_path = os.path.join("template_reports", "psat_2018_analysis_report.xlsx")
act_report_path = os.path.join("template_reports", "act_2018_analysis_report.xlsx")

class Correction_Results(QWidget):
    def __init__(self, file_manager):
        super().__init__()

        #create file manager class
        self.file_manager = file_manager

        #create overall vertical layout
        self.layout = QVBoxLayout(self)

        #extra variable
        self.load = False

        #creating extra layouts
        self.search_layout = QVBoxLayout(self)
        self.select_layout = QHBoxLayout(self)
        self.date_layout = QHBoxLayout(self)
        self.info_layout = QHBoxLayout(self)
        self.sum_layout = QVBoxLayout(self)
        self.detail_layout = QVBoxLayout(self)
        self.sp_layout = QHBoxLayout(self)
        self.save_layout = QVBoxLayout(self)
        self.essay_layout = QHBoxLayout(self)
        self.print_layout = QVBoxLayout(self)
        self.button_layout = QHBoxLayout(self)

        #creating items in search layout
        self.title = QLabel("Search Condition", self)
        self.title.setStyleSheet("color : blue;")

        #creating items in select layout
        self.load_button = QPushButton("Load", self)
        self.load_button.clicked.connect(self.load_sum_table)
        self.save_button = QPushButton("Save", self)
        self.save_button.clicked.connect(self.save_test_to_database)

        #creating items in date layout
        self.date_title = QLabel("Test Date:", self)
        #get current date
        now = datetime.datetime.now()
        self.date_combo = QLabel( str(now.month) + "/" + str(now.day) + "/" + str(now.year) )
        self.change_button = QPushButton("Change Date")
        self.change_button.clicked.connect(self.openCalender)

        # create calendar dialog
        self.cd = Calender_Dialog(self.date_combo)

        #creating items in sum layout
        self.sum_title = QLabel("Summary Information", self)
        self.sum_title.setStyleSheet("color : blue;")
        self.select_check = QCheckBox("select all", self)
        self.select_check.stateChanged.connect(self.select_all_checked )
        self.initSumTable()
        self.sum_table.cellClicked.connect(self.highlight_row)

        #creating items in detail layout
        self.detail_title = QLabel("Detail Information", self)
        self.detail_title.setStyleSheet("color : blue;")
        self.initDetailTable()

        #creating items in save layout
        self.save_title = QLabel("Input Essay Score", self)
        self.save_title.setStyleSheet("color : blue;")

        #creating items in essay layout
        self.essay_title = QLabel("Essay Score:", self)
        self.essay_combo = self.initComboBox(["1", "2", "3", "4", "5", "6"])
        self.essay_save_button = QPushButton("Save Essay Score", self)

        #creating items in print layout
        self.print_title = QLabel("Print Report", self)
        self.print_title.setStyleSheet("color : blue;")

        #creating items in button layout
        self.print_button1 = QPushButton("Print Test\nReport", self)
        self.print_button2 = QPushButton("Generate\nTest Report", self)
        self.print_button2.clicked.connect(self.generate_report)

        #adding to search layout
        self.search_layout.addWidget(self.title)
        self.search_layout.addLayout(self.select_layout)

        #adding to select layout
        self.select_layout.addLayout(self.date_layout)
        self.select_layout.addWidget(self.load_button)
        self.select_layout.addWidget(self.save_button)

        #adding to date layout
        self.date_layout.addWidget(self.date_title)
        self.date_layout.addWidget(self.date_combo)
        self.date_layout.addWidget(self.change_button)

        #adding to info layout
        self.info_layout.addLayout(self.sum_layout)
        self.info_layout.addLayout(self.detail_layout)

        #adding to sum layout
        self.sum_layout.addWidget(self.sum_title)
        self.sum_layout.addWidget(self.select_check)
        self.sum_layout.addWidget(self.sum_table)

        #adding to detail layout
        self.detail_layout.addWidget(self.detail_title)
        self.detail_layout.addWidget(self.detail_table)

        #adding to sp layout
        self.sp_layout.addLayout(self.save_layout)
        self.sp_layout.addLayout(self.print_layout)

        #adding to save layout
        self.save_layout.addWidget(self.save_title)
        self.save_layout.addLayout(self.essay_layout)

        #adding to essay layout
        self.essay_layout.addWidget(self.essay_title)
        self.essay_layout.addWidget(self.essay_combo)
        self.essay_layout.addWidget(self.essay_save_button)

        #adding to print layout
        self.print_layout.addWidget(self.print_title)
        self.print_layout.addLayout(self.button_layout)

        #adding to button layout
        self.button_layout.addWidget(self.print_button1)
        self.button_layout.addWidget(self.print_button2)

        #adding to overall layout
        self.layout.addLayout(self.search_layout)
        self.layout.addLayout(self.info_layout)
        self.layout.addLayout(self.sp_layout)

        self.setLayout(self.layout)

    def initSumTable(self):
        # create table
        self.sum_table = QTableWidget()

        self.row = 0


        self.col = 10

        row = self.row
        col = self.col

        self.sum_table.setRowCount(row)
        self.sum_table.setFocusPolicy(Qt.StrongFocus)
        self.sum_table.setColumnCount(col)

        table_labels = ["Date","Name" , "Student ID", "Subject", "Category", "Number", "Section", "Correct/Total" ,"%", "Select"]

        header = self.sum_table.horizontalHeader()
        #set the table column labels
        for i in range(0, len(table_labels)):
            self.sum_table.setHorizontalHeaderItem(i, QTableWidgetItem(table_labels[i]))
            if i == 0:
                header.setSectionResizeMode(i, QtWidgets.QHeaderView.Stretch)
            else:
                header.setSectionResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)

        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        #create comboxbox for each cell
        for i in range(0, row):
            for j in range(0, col):
                if j < col - 1:
                    self.sum_table.setItem(i, j, QTableWidgetItem(""))
                else:
                    checkbox = QCheckBox("", self)
                    self.sum_table.setCellWidget(i, j, checkbox)

    def initDetailTable(self):
        # create table
        self.detail_table = QTableWidget()

        self.d_row = 0
        self.d_col = 5

        row = self.d_row
        col = self.d_col

        self.detail_table.setRowCount(row)
        self.detail_table.setColumnCount(col)

        table_labels = ["Answer Type" ,"Question Type","Difficulty" , "Answer", "Status" ]

        header = self.detail_table.horizontalHeader()
        #set the table column labels
        for i in range(0, len(table_labels)):
            self.detail_table.setHorizontalHeaderItem(i, QTableWidgetItem(table_labels[i]))
            if i == 0:
                header.setSectionResizeMode(i, QtWidgets.QHeaderView.Stretch)
            else:
                header.setSectionResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)

        #create comboxbox for each cell
        for i in range(0, row):
            for j in range(0, col):
                self.detail_table.setItem(i, j, QTableWidgetItem(""))

    def initComboBox(self, items):
        combo = QComboBox(self)
        for item in items:
            combo.addItem(item)
        return combo

    def openCalender(self):
        self.cd.show()

    def select_all_checked(self):

        for i in range(0, self.row):
            self.sum_table.cellWidget(i, self.col - 1).setChecked(self.select_check.isChecked())

    def load_sum_table(self):
        data = self.file_manager.get_log( self.date_combo.text() )
        if len(data) == 0:
            self.clear_table_rows()
            buttonReply = QMessageBox.warning(self, 'Warning', "Date you selected is empty!",
                                              QMessageBox.Ok, QMessageBox.Ok)
            self.load = False
        else:
            self.load = True
            self.clear_table_rows()
            self.row = len(data)
            self.col = 10


            for i in range(0, self.row):
                self.sum_table.insertRow(i)
                first_name = self.file_manager.get_user_info(data[i][0])[0]
                last_name = self.file_manager.get_user_info(data[i][0])[1]
                name = first_name + " " + last_name
                for j in range(0, self.col):
                    if j == 0:
                        self.sum_table.setItem(i, j, QTableWidgetItem( self.date_combo.text() ))
                    elif j == self.col - 1:
                        checkbox = QCheckBox("", self)
                        self.sum_table.setCellWidget(i, j, checkbox)
                    elif j > 1:
                        self.sum_table.setItem(i, j, QTableWidgetItem( str(data[i][j-2] ) ))
                    else:
                        self.sum_table.setItem(i, j, QTableWidgetItem(name) )

            buttonReply = QMessageBox.information(self, 'Notification', "Loaded successfully!",
                                                  QMessageBox.Ok, QMessageBox.Ok)



    def clear_table_rows(self):
        self.row = 0
        while self.sum_table.rowCount() > 0:
            self.sum_table.removeRow(0)

        self.d_row = 0
        while self.detail_table.rowCount() > 0:
            self.detail_table.removeRow(0)

    def clear_detail_rows(self):
        self.d_row = 0
        while self.detail_table.rowCount() > 0:
            self.detail_table.removeRow(0)

    def highlight_row(self, row, col):
        self.sum_table.selectRow(row)
        self.show_row_detail(row)

    def show_row_detail(self, row):
        self.clear_detail_rows()
        data = self.file_manager.get_log(self.date_combo.text())[row][7]

        #print("data:" , data)
        self.d_row = len(data)
        self.d_col = len(data[0])

        for i in range(0, self.d_row):
            self.detail_table.insertRow(i)
            for j in range(0, self.d_col):
                if j == self.d_col - 1:
                    self.detail_table.setItem(i, j, QTableWidgetItem("") )
                    if data[i][j] == "C":
                        self.detail_table.item(i, j).setBackground(QtGui.QColor(0, 255, 0))
                    elif data[i][j] == "W":
                        self.detail_table.item(i, j).setBackground(QtGui.QColor(255, 0, 0))
                    else:
                        self.detail_table.item(i, j).setBackground(QtGui.QColor(255, 165, 0))
                else:
                    self.detail_table.setItem(i, j, QTableWidgetItem( data[i][j] ) )

    def save_test_to_database(self):
        if self.row == 0:
            buttonReply = QMessageBox.warning(self, 'Warning', "Nothing is selected!",
                                              QMessageBox.Ok, QMessageBox.Ok)
            return

        data = self.file_manager.get_log(self.date_combo.text())

        for i in range(0, self.row):
            if self.sum_table.cellWidget(i, self.col - 1).isChecked():
                id = self.sum_table.item(i, 2 ).text()
                self.file_manager.save_test_to_student( id, self.date_combo.text(), data[i] )

        self.file_manager.save_file("cr")
        buttonReply = QMessageBox.information(self, 'Notification', "Test results were saved to Student Record successfully!",
                                              QMessageBox.Ok, QMessageBox.Ok)

    def generate_report(self):
        if self.load == False:
            buttonReply = QMessageBox.warning(self, 'Warning', "Load tests before generating report!",
                                              QMessageBox.Ok, QMessageBox.Ok)
            return

        report_dict = {}
        score_dict = {}

        date = self.sum_table.item(0, 0).text()
        for i in range(0, self.row):
            if self.sum_table.cellWidget(i, self.col - 1).isChecked():

                id = self.sum_table.item(i, 2 ).text()

                subject = self.sum_table.item(i, 3 ).text()
                category = self.sum_table.item(i, 4 ).text()
                number = self.sum_table.item(i, 5 ).text()
                test_num = subject + "-" + category + "-" + number

                section = self.sum_table.item(i, 6 ).text()
                result = self.sum_table.item(i, 7 ).text().split("/")
                cor_num = int(result[0])
                total_num = int(result[1])

                if id not in report_dict.keys():
                    report_dict[id] = {}
                    score_dict[id] = {}

                if test_num not in report_dict[id].keys():
                    report_dict[id][test_num] = []
                    score_dict[id][test_num] = {}

                section_name = section
                if section in report_dict[id][test_num]:
                    section_name += "1"

                report_dict[id][test_num].append(section_name)
                report_dict[id][test_num].sort()

                score_dict[id][test_num][section_name] =  [cor_num, total_num]


        if len(report_dict.keys()) == 0:
            buttonReply = QMessageBox.warning(self, 'Warning', "Select at least one test!",
                                              QMessageBox.Ok, QMessageBox.Ok)
            return

        cur_dir = os.getcwd()
        pathlib.Path(cur_dir + '/score_reports').mkdir(parents=True, exist_ok=True)

        now = datetime.datetime.now()
        cur_date = str(now.month) + "-" + str(now.day) + "-" + str(now.year)
        pathlib.Path(cur_dir + '/score_reports/' + cur_date).mkdir(parents=True, exist_ok=True)

        for id in report_dict.keys():
            user_info = self.file_manager.get_user_info(id)
            test_info = self.file_manager.get_test_info(id)



            for test_num in report_dict[id].keys():
                source = cur_dir
                mode = None
                if "SAT I" in test_num:
                    mode = "SAT I"
                    source = os.path.join(source, sat_report_path)
                elif "PSAT" in test_num:
                    mode = "PSAT"
                    source = os.path.join(source, psat_report_path)
                elif "ACT" in test_num:
                    mode = "ACT"
                    source = os.path.join(source, act_report_path)

                full_name = user_info[0] + user_info[1]
                grade = user_info[2]
                student_id = user_info[3]

                filename = user_info[0] + "_" + user_info[1] + ":" + test_num + ":Report"
                filename = "".join(filename.split())

                #dest = cur_dir + '/score_reports/' + cur_date + '/' + filename + ".xlsx"

                filename = filename.replace(':', '_')

                dest = os.path.join(cur_dir, 'score_reports', cur_date,  filename + ".xlsx" )
                #dest = dest.replace(":", "_")

                copyfile(source, dest)


                wb = load_workbook(filename=dest)

                if mode == "SAT I" or mode == "PSAT":
                    count_table = [0,0,0] #math, reading, writing
                    for section in report_dict[id][test_num]:
                        if "Math" in section:
                            count_table[0] += 1
                        elif "Reading" in section:
                            count_table[1] += 1
                        elif "Writing" in section:
                            count_table[2] += 1
                    idx = 3
                    for i in range(0, len(count_table)):
                        if i == 0 and count_table[i] == 0:
                            ws = wb["Math"]
                            wb.remove(ws)
                        elif i == 1 and count_table[i] == 0:
                            ws = wb["Reading"]
                            wb.remove(ws)
                        elif i == 2 and count_table[i] == 0:
                            ws = wb["Writing"]
                            wb.remove(ws)

                        if i == 0 and count_table[i] > 2: #math
                            ws = wb["Math"]
                            while count_table[i] > 2:
                                wb.copy_worksheet(ws)
                                count_table[i] -= 1
                        elif i == 1 and count_table[i] > 1:
                            ws = wb["Reading"]
                            while count_table[i] > 1:
                                wb.copy_worksheet(ws)
                                count_table[i] -= 1
                        elif i == 2 and count_table[i] > 1:
                            ws = wb["Writing"]
                            while count_table[i] > 1:
                                wb.copy_worksheet(ws)
                                count_table[i] -= 1


                pos = None
                score_map = {}

                for name in wb.sheetnames:

                    ws = wb[name]
                    img = Image('icon.png')
                    ws.add_image(img, 'A2')

                    prefix = ""
                    name_pos = "2"
                    grade_pos = "3"
                    id_pos = "4"
                    exam_pos = "5"
                    date_pos = "6"

                    test_type = test_num.split("-")
                    exam = test_type[1] + "-" + test_type[2]
                    category = test_type[1]
                    number = test_type[2]

                    if mode == "SAT I" or mode == "PSAT":

                        if name == "Summary":
                            prefix = "Q"
                            if "Math" not in score_map.keys():
                                score_map["Math"] = [0, 0]

                            ws["E10"] = score_map["Math"][0]
                            ws["I10"] = score_map["Math"][1]
                            ws["M10"] = str(score_map["Math"][1]) + "/" + "800"

                            if "Reading" not in score_map.keys():
                                score_map["Reading"] = [0, 0]
                            ws["E11"] = score_map["Reading"][0]
                            ws["I11"] = score_map["Reading"][1]

                            if "Writing" not in score_map.keys():
                                score_map["Writing"] = [0, 0]
                            ws["E12"] = score_map["Writing"][0]
                            ws["I12"] = score_map["Writing"][1]
                            ws["M11"] = str( score_map["Reading"][1] + score_map["Writing"][1] ) + "/" + "800"

                            ws["Q10"] = str( score_map["Math"][1] + score_map["Reading"][1] + score_map["Writing"][1] ) + "/" + "1600"

                            history_map = {}
                            key_list = ["Reading", "Writing", "Math"]

                            for x in range(0, len(key_list) ):

                                for i in range(0, len(score_map[key_list[x]][2]) ):
                                    if score_map[key_list[x]][2][i][0] not in history_map.keys():
                                        history_map[score_map[key_list[x]][2][i][0]] = {} #math, reading, writng
                                    if score_map[key_list[x]][2][i][1] not in history_map[score_map[key_list[x]][2][i][0]].keys():
                                        history_map[score_map[key_list[x]][2][i][0]][score_map[key_list[x]][2][i][1]] = [0,0,0]
                                    history_map[score_map[key_list[x]][2][i][0]][score_map[key_list[x]][2][i][1]][x] += score_map[key_list[x]][2][i][2]

                            loc = 27
                            prev_score = None
                            average = [0, 0, 0, 0]
                            count = 0
                            for key in history_map.keys():
                                for test_id in history_map[key].keys():
                                    ws["A" + str(loc)] = loc - 26
                                    ws["B" + str(loc)] = key + "," + test_id
                                    ws["D" + str(loc)] = history_map[key][test_id][2]
                                    average[0] += history_map[key][test_id][2]
                                    ws["F" + str(loc)] = history_map[key][test_id][0] + history_map[key][test_id][1]
                                    average[1] += history_map[key][test_id][0] + history_map[key][test_id][1]
                                    ws["I" + str(loc)] = history_map[key][test_id][0] + history_map[key][test_id][1] + history_map[key][test_id][2]
                                    total = history_map[key][test_id][0] + history_map[key][test_id][1] + history_map[key][test_id][2]
                                    average[2] += total

                                    change = None
                                    if prev_score == None:
                                        prev_score = total
                                        change = 0
                                    else:
                                        change = total - prev_score
                                    average[3] += change
                                    ws["K" + str(loc)] = change
                                    loc += 1
                                    count += 1



                            ws["D51"] = int(average[0]/count)
                            ws["F51"] = int(average[1]/count)
                            ws["I51"] = int(average[2]/count)
                            ws["K51"] = int(average[3]/count)
                            #print("map: ", history_map)


                        else:
                            dates = list(test_info.keys())
                            dates.sort()


                            if name == "Reading":
                                cor_count = score_dict[id][test_num][name][0]
                                total_count = score_dict[id][test_num][name][1]
                                blank_count = 0
                                answers = self.file_manager.get_tm_table_data(mode, category, number, "Reading")
                                raw_table = self.file_manager.get_rt(mode, "Reading")


                                pos = [i for i,x in enumerate(test_info[date]) if (x[3] == "Reading" and
                                                                                   (x[1] + "-" + x[2]) == exam
                                                                                   )]

                                ws["A10"] = 1
                                ws["J10"] = cor_count
                                ws["K10"] = total_count - cor_count
                                ws["M10"] = cor_count
                                ws["P10"] = raw_table[int(cor_count)] * 10

                                score_map["Reading"] = [ raw_table[int(cor_count)], raw_table[int(cor_count)] * 10, [] ]
                                #score_map["Reading"] = []

                                answer_type = {}
                                level_type = {}

                                for i in range(1, total_count + 1):
                                    ws["B" + str(9 + i)] = i # no. cell
                                    ws["C" + str(9 + i)] = answers[i-1][3]
                                    ws["D" + str(9 + i)] = test_info[date][pos[0]][6][i-1][3]  # answer cell
                                    if test_info[date][pos[0]][6][i-1][3] == " ":
                                        ws["D" + str(9 + i)] = "(blank)"
                                        blank_count += 1
                                    if test_info[date][pos[0]][6][i - 1][4] == "W" or test_info[date][pos[0]][6][i - 1][
                                        4] == "B":
                                        my_red = openpyxl.styles.colors.Color(rgb='E88787')
                                        ws["B" + str(9 + i)].fill = openpyxl.styles.fills.PatternFill(
                                            patternType='solid', fgColor=my_red)
                                    ws["E" + str(9 + i)] = test_info[date][pos[0]][6][i - 1][2]  # difficulty cell
                                    ws["H" + str(9 + i)] = test_info[date][pos[0]][6][i - 1][1]  # type cell

                                    question_type = test_info[date][pos[0]][6][i - 1][1]
                                    diff_type = test_info[date][pos[0]][6][i - 1][2]

                                    #saving question type and level type
                                    if question_type not in answer_type.keys():
                                        answer_type[question_type] = [0, 0]
                                    if diff_type not in level_type.keys():
                                        level_type[diff_type] = [0, 0]

                                    if test_info[date][pos[0]][6][i-1][4] == "C":
                                        answer_type[question_type][0] += 1
                                        level_type[diff_type][0] += 1
                                    answer_type[question_type][1] += 1
                                    level_type[diff_type][1] += 1
                                    #saving difficulty type

                                row_id = 52
                                for key in answer_type.keys():
                                    ws["J" + str(row_id)] = key
                                    ws["O" + str(row_id)] = answer_type[key][1]
                                    ws["P" + str(row_id)] = answer_type[key][0]
                                    ws["Q" + str(row_id)] = self.get_percent(answer_type[key][0], answer_type[key][1])
                                    row_id += 1

                                row_id = 65
                                for key in level_type.keys():
                                    ws["J" + str(row_id)] = key
                                    ws["M" + str(row_id)] = level_type[key][1]
                                    ws["O" + str(row_id)] = level_type[key][0]
                                    ws["Q" + str(row_id)] = self.get_percent(level_type[key][0], level_type[key][1])
                                    row_id += 1

                                insert_count = 1
                                row_id = 14
                                prev_score = None
                                average = [0,0]
                                count = 0
                                for item in dates:
                                    for item1 in test_info[item]:
                                        if item1[3] == "Reading" and mode == item1[0]:
                                            ws["J" + str(row_id)] = insert_count
                                            insert_count += 1
                                            ws["K" + str(row_id)] = (item + "," + item1[1] + "-" + item1[2])

                                            score = item1[4].split("/")[0]
                                            score = raw_table[int(score)] * 10
                                            score_map["Reading"][2].append([ item ,(item1[1] + "-" + item1[2]), score ])
                                            ws["M" + str(row_id)] = score

                                            change = None
                                            if prev_score == None:
                                                prev_score = score
                                                change = 0
                                            else:
                                                change = score - prev_score
                                                prev_score = score
                                            average[0] += score
                                            average[1] += change
                                            count += 1
                                            ws["P" + str(row_id)] = change
                                            row_id += 1


                                ws["L10"] = blank_count
                                ws["M38"] = int(average[0]/count)
                                ws["P38"] = int(average[1]/count)

                            elif name == "Writing":
                                section_num = 2
                                cor_count = score_dict[id][test_num]["Writing and Language"][0]
                                total_count = score_dict[id][test_num]["Writing and Language"][1]
                                blank_count = 0
                                answers = self.file_manager.get_tm_table_data(mode, category, number, "Writing and Language")
                                raw_table = self.file_manager.get_rt(mode, "Writing")

                                pos = [i for i, x in enumerate(test_info[date]) if ( "Writing" in x[3] and
                                                                                    (x[1] + "-" + x[2]) == exam
                                                                                    )]

                                ws["A10"] = 2
                                ws["J10"] = cor_count
                                ws["K10"] = total_count - cor_count
                                ws["M10"] = cor_count
                                ws["P10"] = raw_table[int(cor_count)] * 10
                                score_map["Writing"] = [raw_table[int(cor_count)] , raw_table[int(cor_count)] * 10, [] ]

                                answer_type = {}
                                level_type = {}

                                for i in range(1, total_count + 1):
                                    ws["B" + str(9 + i)] = i  # no. cell
                                    ws["C" + str(9 + i)] = answers[i - 1][3]
                                    ws["D" + str(9 + i)] = test_info[date][pos[0]][6][i - 1][3]  # answer cell
                                    if test_info[date][pos[0]][6][i-1][3] == " ":
                                        ws["D" + str(9 + i)] = "(blank)"
                                        blank_count += 1
                                    if test_info[date][pos[0]][6][i - 1][4] == "W" or test_info[date][pos[0]][6][i - 1][
                                        4] == "B":
                                        my_red = openpyxl.styles.colors.Color(rgb='E88787')
                                        ws["B" + str(9 + i)].fill = openpyxl.styles.fills.PatternFill(
                                            patternType='solid', fgColor=my_red)
                                    ws["E" + str(9 + i)] = test_info[date][pos[0]][6][i - 1][2]  # difficulty cell
                                    ws["H" + str(9 + i)] = test_info[date][pos[0]][6][i - 1][1]  # type cell

                                    question_type = test_info[date][pos[0]][6][i - 1][1]
                                    diff_type = test_info[date][pos[0]][6][i - 1][2]

                                    # saving question type and level type
                                    if question_type not in answer_type.keys():
                                        answer_type[question_type] = [0, 0]
                                    if diff_type not in level_type.keys():
                                        level_type[diff_type] = [0, 0]

                                    if test_info[date][pos[0]][6][i - 1][4] == "C":
                                        answer_type[question_type][0] += 1
                                        level_type[diff_type][0] += 1
                                    answer_type[question_type][1] += 1
                                    level_type[diff_type][1] += 1

                                row_id = 52
                                for key in answer_type.keys():
                                    ws["J" + str(row_id)] = key
                                    ws["O" + str(row_id)] = answer_type[key][1]
                                    ws["P" + str(row_id)] = answer_type[key][0]
                                    ws["Q" + str(row_id)] = self.get_percent(answer_type[key][0], answer_type[key][1])
                                    row_id += 1

                                row_id = 65
                                for key in level_type.keys():
                                    ws["J" + str(row_id)] = key
                                    ws["M" + str(row_id)] = level_type[key][1]
                                    ws["O" + str(row_id)] = level_type[key][0]
                                    ws["Q" + str(row_id)] = self.get_percent(level_type[key][0], level_type[key][1])
                                    row_id += 1

                                insert_count = 1
                                row_id = 14
                                prev_score = None
                                average = [0, 0]
                                count = 0

                                for item in dates:
                                    for item1 in test_info[item]:
                                        if "Writing" in item1[3] and mode == item1[0]:
                                            ws["J" + str(row_id)] = insert_count
                                            insert_count += 1
                                            ws["K" + str(row_id)] = (item + "," + item1[1] + "-" + item1[2])
                                            score = item1[4].split("/")[0]
                                            score = raw_table[int(score)] * 10
                                            score_map["Writing"][2].append( [item, item1[1] + "-" + item1[2], raw_table[int(cor_count)] * 10] )
                                            ws["M" + str(row_id)] = score
                                            change = None
                                            if prev_score == None:
                                                prev_score = score
                                                change = 0
                                            else:
                                                change = score - prev_score
                                                prev_score = score
                                            average[0] += score
                                            average[1] += change
                                            count += 1
                                            ws["P" + str(row_id)] = change
                                            row_id += 1

                                ws["L10"] = blank_count
                                ws["M38"] = int(average[0] / count)
                                ws["P38"] = int(average[1] / count)




                            elif name == "Math":

                                cor_count = 0
                                cor_count1 = 0
                                total_count = 0
                                total_count1 = 0

                                math_no = False
                                math = False

                                if "Math No Calculator" in score_dict[id][test_num].keys():
                                    cor_count = score_dict[id][test_num]["Math No Calculator"][0]
                                    total_count = score_dict[id][test_num]["Math No Calculator"][1]
                                    math_no = True

                                if "Math Calculator" in score_dict[id][test_num].keys():
                                    cor_count1 = score_dict[id][test_num]["Math Calculator"][0]
                                    total_count1 = score_dict[id][test_num]["Math Calculator"][1]
                                    math = True

                                blank_count = 0

                                section_num = 3

                                raw_table = self.file_manager.get_rt(mode, "Math")
                                answer_type = {}
                                level_type = {}

                                score_map["Math"] = [cor_count1 + cor_count, raw_table[int(cor_count + cor_count1)], [] ]

                                if math_no == True:
                                    answers = self.file_manager.get_tm_table_data(mode, category, number, "Math No Calculator")

                                    pos = [i for i, x in enumerate(test_info[date]) if (x[3] == "Math No Calculator" and
                                                                                        (x[1] + "-" + x[2]) == exam
                                                                                        )]
                                    for i in range(1, total_count + 1):
                                        ws["B" + str(9 + i)] = i  # no. cell
                                        ws["C" + str(9 + i)] = answers[i - 1][3]
                                        ws["D" + str(9 + i)] = test_info[date][pos[0]][6][i - 1][3]  # answer cell
                                        if test_info[date][pos[0]][6][i - 1][3] == " ":
                                            ws["D" + str(9 + i)] = "(blank)"
                                            blank_count += 1

                                        if test_info[date][pos[0]][6][i - 1][4] == "W" or test_info[date][pos[0]][6][i - 1][4] == "B":

                                            my_red = openpyxl.styles.colors.Color(rgb='E88787')
                                            ws["B" + str(9 + i)].fill = openpyxl.styles.fills.PatternFill(
                                                patternType='solid', fgColor=my_red)
                                        ws["E" + str(9 + i)] = test_info[date][pos[0]][6][i - 1][2]  # difficulty cell
                                        ws["H" + str(9 + i)] = test_info[date][pos[0]][6][i - 1][1]  # type cell


                                        question_type = test_info[date][pos[0]][6][i - 1][1]
                                        diff_type = test_info[date][pos[0]][6][i - 1][2]

                                        # saving question type and level type
                                        if question_type not in answer_type.keys():
                                            answer_type[question_type] = [0, 0]
                                        if diff_type not in level_type.keys():
                                            level_type[diff_type] = [0, 0]

                                        if test_info[date][pos[0]][6][i - 1][4] == "C":
                                            answer_type[question_type][0] += 1
                                            level_type[diff_type][0] += 1
                                        answer_type[question_type][1] += 1
                                        level_type[diff_type][1] += 1

                                ws["A10"] = 3
                                ws["A30"] = 4
                                ws["J10"] = cor_count + cor_count1
                                ws["K10"] = total_count + total_count1 - cor_count - cor_count1
                                ws["M10"] = cor_count + cor_count1
                                ws["P10"] = raw_table[int(cor_count + cor_count1)]

                                if math == True:
                                    answers = self.file_manager.get_tm_table_data(mode, category, number,
                                                                                  "Math Calculator")
                                    pos = [i for i, x in enumerate(test_info[date]) if (x[3] == "Math Calculator" and
                                                                                        (x[1] + "-" + x[2]) == exam
                                                                                        )]

                                    for i in range(1, total_count1 + 1):
                                        ws["B" + str(9 + i + 20)] = i  # no. cell
                                        ws["C" + str(9 + i + 20)] = answers[i - 1][3]
                                        ws["D" + str(9 + i + 20)] = test_info[date][pos[0]][6][i - 1][3]  # answer cell
                                        if test_info[date][pos[0]][6][i - 1][3] == " ":
                                            ws["D" + str(9 + i + 20)] = "(blank)"
                                            blank_count += 1
                                        color = openpyxl.styles.colors.Color(rgb='E88787')
                                        if test_info[date][pos[0]][6][i - 1][4] == "W" or test_info[date][pos[0]][6][i - 1][4] == "B":
                                            ws["B" + str(9 + i + 20)].fill = openpyxl.styles.fills.PatternFill(
                                                patternType='solid', fgColor=color)


                                        ws["E" + str(9 + i + 20)] = test_info[date][pos[0]][6][i - 1][2]  # difficulty cell
                                        ws["H" + str(9 + i + 20)] = test_info[date][pos[0]][6][i - 1][1]  # type cell

                                        question_type = test_info[date][pos[0]][6][i - 1][1]
                                        diff_type = test_info[date][pos[0]][6][i - 1][2]

                                        # saving question type and level type
                                        if question_type not in answer_type.keys():
                                            answer_type[question_type] = [0, 0]
                                        if diff_type not in level_type.keys():
                                            level_type[diff_type] = [0, 0]

                                        if test_info[date][pos[0]][6][i - 1][4] == "C":
                                            answer_type[question_type][0] += 1
                                            level_type[diff_type][0] += 1
                                        answer_type[question_type][1] += 1
                                        level_type[diff_type][1] += 1

                                row_id = 52
                                for key in answer_type.keys():
                                    ws["J" + str(row_id)] = key
                                    ws["O" + str(row_id)] = answer_type[key][1]
                                    ws["P" + str(row_id)] = answer_type[key][0]
                                    ws["Q" + str(row_id)] = self.get_percent(answer_type[key][0], answer_type[key][1])
                                    row_id += 1

                                row_id = 65
                                for key in level_type.keys():
                                    ws["J" + str(row_id)] = key
                                    ws["M" + str(row_id)] = level_type[key][1]
                                    ws["O" + str(row_id)] = level_type[key][0]
                                    ws["Q" + str(row_id)] = self.get_percent(level_type[key][0], level_type[key][1])
                                    row_id += 1



                                match_map = {}
                                for item in dates:
                                    for item1 in test_info[item]:
                                        if "Math" in item1[3] and mode == item1[0]:
                                            score = item1[4].split("/")[0]

                                            if (item + "," + item1[1] + "-" + item1[2]) not in match_map.keys():
                                                match_map[(item + "," + item1[1] + "-" + item1[2])] = int(score)

                                            else: #first math section was discovered
                                                match_map[(item + "," + item1[1] + "-" + item1[2])] += int(score)

                                dates = list( match_map.keys() )
                                dates.sort()

                                insert_count = 1
                                row_id = 14
                                prev_score = None
                                average = [0, 0]
                                count = 0
                                for item in dates:
                                    ws["J" + str(row_id)] = insert_count
                                    insert_count += 1
                                    ws["K" + str(row_id)] = item
                                    score = raw_table[match_map[item]]
                                    ws["M" + str(row_id)] = score
                                    items = item.split(",")
                                    score_map["Math"][2].append( [items[0], items[1],score] )
                                    change = None
                                    if prev_score == None:
                                        prev_score = score
                                        change = 0
                                    else:
                                        change = score - prev_score
                                        prev_score = score
                                    average[0] += score
                                    average[1] += change
                                    count += 1
                                    ws["P" + str(row_id)] = change
                                    row_id += 1

                                ws["L10"] = blank_count
                                ws["M38"] = int(average[0] / count)
                                ws["P38"] = int(average[1] / count)








                            prefix = "O"

                    elif mode == "ACT":

                        if name == "Summary":
                            prefix = "Q"
                            total = 0
                            t_c = 0
                            if "English" not in score_map.keys():
                                score_map["English"] = [0, 0]
                            else:
                                t_c += 1

                            ws["E10"] = score_map["English"][0]
                            ws["I10"] = score_map["English"][1]
                            total += score_map["English"][1]
                            ws["M10"] = str(score_map["English"][1]) + "/" + "36"

                            if "Math" not in score_map.keys():
                                score_map["Math"] = [0, 0]
                            else:
                                t_c += 1

                            ws["E11"] = score_map["Math"][0]
                            ws["I11"] = score_map["Math"][1]
                            total += score_map["Math"][1]
                            ws["M11"] = str(score_map["Math"][1]) + "/" + "36"

                            if "Reading" not in score_map.keys():
                                score_map["Reading"] = [0, 0]
                            else:
                                t_c += 1

                            ws["E12"] = score_map["Reading"][0]
                            ws["I12"] = score_map["Reading"][1]
                            total += score_map["Reading"][1]
                            ws["M12"] = str(score_map["Reading"][1]) + "/" + "36"

                            if "Science" not in score_map.keys():
                                score_map["Science"] = [0, 0]
                            else:
                                t_c += 1

                            ws["E13"] = score_map["Science"][0]
                            ws["I13"] = score_map["Science"][1]
                            total += score_map["Science"][1]
                            ws["M13"] = str(score_map["Science"][1]) + "/" + "36"


                            ws["Q10"] = str( int(total/t_c) ) + "/" + "36"

                            history_map = {}
                            key_list = ["English", "Math", "Reading", "Science"]

                            for x in range(0, len(key_list) ):

                                for i in range(0, len(score_map[key_list[x]][2]) ):
                                    if score_map[key_list[x]][2][i][0] not in history_map.keys():
                                        history_map[score_map[key_list[x]][2][i][0]] = {}
                                    if score_map[key_list[x]][2][i][1] not in history_map[score_map[key_list[x]][2][i][0]].keys():
                                        history_map[score_map[key_list[x]][2][i][0]][score_map[key_list[x]][2][i][1]] = [0,0,0,0]
                                    #english, math, reading, science
                                    history_map[score_map[key_list[x]][2][i][0]][score_map[key_list[x]][2][i][1]][x] += score_map[key_list[x]][2][i][2]

                            loc = 27
                            prev_score = None
                            average = [0, 0, 0, 0, 0, 0] #english, math, reading, science
                            count = 0
                            for key in history_map.keys():
                                for test_id in history_map[key].keys():
                                    ws["A" + str(loc)] = loc - 26
                                    ws["B" + str(loc)] = key + "," + test_id
                                    ws["D" + str(loc)] = history_map[key][test_id][0]
                                    average[0] += history_map[key][test_id][0]
                                    ws["F" + str(loc)] = history_map[key][test_id][1]
                                    average[1] += history_map[key][test_id][1]
                                    ws["H" + str(loc)] = history_map[key][test_id][2]
                                    average[2] += history_map[key][test_id][2]
                                    ws["J" + str(loc)] = history_map[key][test_id][3]
                                    average[3] += history_map[key][test_id][3]


                                    ws["L" + str(loc)] = int( (history_map[key][test_id][0] + history_map[key][test_id][1] + history_map[key][test_id][2] + history_map[key][test_id][3] )/4 )
                                    total = history_map[key][test_id][0] + history_map[key][test_id][1] + history_map[key][test_id][2] + history_map[key][test_id][3]
                                    average[4] += total

                                    change = None
                                    if prev_score == None:
                                        prev_score = total
                                        change = 0
                                    else:
                                        change = total - prev_score
                                    average[5] += change
                                    ws["N" + str(loc)] = change
                                    loc += 1
                                    count += 1



                            ws["D51"] = int(average[0]/count)
                            ws["F51"] = int(average[1]/count)
                            ws["H51"] = int(average[2]/count)
                            ws["J51"] = int(average[3]/count)
                            ws["L51"] = int(average[4] / count)
                            ws["N51"] = int(average[5] / count)
                            #print("map: ", history_map)


                        else:
                            dates = list(test_info.keys())
                            dates.sort()

                            cor_count = score_dict[id][test_num][name][0]
                            total_count = score_dict[id][test_num][name][1]
                            blank_count = 0
                            answers = self.file_manager.get_tm_table_data(mode, category, number, name)
                            raw_table = self.file_manager.get_rt(mode, name)


                            pos = [i for i,x in enumerate(test_info[date]) if (x[3] == name and
                                                                               (x[1] + "-" + x[2]) == exam
                                                                               )]

                            ws["A10"] = 1
                            ws["J10"] = cor_count
                            ws["K10"] = total_count - cor_count
                            ws["M10"] = raw_table[int(cor_count)]
                            ws["P10"] = raw_table[int(cor_count)]

                            score_map[name] = [ raw_table[int(cor_count)], raw_table[int(cor_count)], [] ]

                            answer_type = {}
                            level_type = {}

                            for i in range(1, total_count + 1):
                                ws["B" + str(9 + i)] = i # no. cell
                                ws["C" + str(9 + i)] = answers[i-1][3]
                                ws["D" + str(9 + i)] = test_info[date][pos[0]][6][i-1][3]  # answer cell
                                if test_info[date][pos[0]][6][i-1][3] == " ":
                                    ws["D" + str(9 + i)] = "(blank)"
                                    blank_count += 1
                                if test_info[date][pos[0]][6][i - 1][4] == "W" or test_info[date][pos[0]][6][i - 1][
                                    4] == "B":
                                    my_red = openpyxl.styles.colors.Color(rgb='E88787')
                                    ws["B" + str(9 + i)].fill = openpyxl.styles.fills.PatternFill(
                                        patternType='solid', fgColor=my_red)
                                ws["E" + str(9 + i)] = test_info[date][pos[0]][6][i - 1][2]  # difficulty cell
                                ws["H" + str(9 + i)] = test_info[date][pos[0]][6][i - 1][1]  # type cell

                                question_type = test_info[date][pos[0]][6][i - 1][1]
                                diff_type = test_info[date][pos[0]][6][i - 1][2]

                                #saving question type and level type
                                if question_type not in answer_type.keys():
                                    answer_type[question_type] = [0, 0]
                                if diff_type not in level_type.keys():
                                    level_type[diff_type] = [0, 0]

                                if test_info[date][pos[0]][6][i-1][4] == "C":
                                    answer_type[question_type][0] += 1
                                    level_type[diff_type][0] += 1
                                answer_type[question_type][1] += 1
                                level_type[diff_type][1] += 1
                                #saving difficulty type

                            row_id = 52
                            for key in answer_type.keys():
                                ws["J" + str(row_id)] = key
                                ws["O" + str(row_id)] = answer_type[key][1]
                                ws["P" + str(row_id)] = answer_type[key][0]
                                ws["Q" + str(row_id)] = self.get_percent(answer_type[key][0], answer_type[key][1])
                                row_id += 1

                            row_id = 65
                            for key in level_type.keys():
                                ws["J" + str(row_id)] = key
                                ws["M" + str(row_id)] = level_type[key][1]
                                ws["O" + str(row_id)] = level_type[key][0]
                                ws["Q" + str(row_id)] = self.get_percent(level_type[key][0], level_type[key][1])
                                row_id += 1

                            insert_count = 1
                            row_id = 14
                            prev_score = None
                            average = [0,0]
                            count = 0

                            for item in dates:
                                for item1 in test_info[item]:
                                    if item1[3] == name and mode == item1[0]:
                                        ws["J" + str(row_id)] = insert_count
                                        insert_count += 1
                                        ws["K" + str(row_id)] = (item + "," + item1[1] + "-" + item1[2])

                                        score = item1[4].split("/")[0]
                                        score = raw_table[int(score)]
                                        score_map[name][2].append([ item ,(item1[1] + "-" + item1[2]), score ])
                                        ws["M" + str(row_id)] = score

                                        change = None
                                        if prev_score == None:
                                            prev_score = score
                                            change = 0
                                        else:
                                            change = score - prev_score
                                            prev_score = score
                                        average[0] += score
                                        average[1] += change
                                        count += 1
                                        ws["P" + str(row_id)] = change
                                        row_id += 1


                            ws["L10"] = blank_count
                            ws["M38"] = int(average[0]/count)
                            ws["P38"] = int(average[1]/count)

                            prefix = "O"



                    ws[prefix + name_pos] = full_name
                    ws[prefix + grade_pos] = grade
                    ws[prefix + id_pos] = id
                    ws[prefix + exam_pos] = exam
                    ws[prefix + date_pos] = date




                wb.save(filename=dest)
                self.open_file(dest)


            #change it so the file does not save with same name


    def open_file(self, path1):
        if platform.system() == "Windows":
            os.startfile(path1)
        elif platform.system() == "Darwin":
            subprocess.Popen(["open", path1])
        else:
            subprocess.Popen(["xdg-open", path1])

    def get_percent(self, correct, total_count):
        return int( round(correct/float(total_count), 2) * 100)





