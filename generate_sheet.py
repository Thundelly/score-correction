import datetime
import pathlib
import os
from shutil import *
from openpyxl import *
from openpyxl.drawing.image import Image
from openpyxl import styles
import openpyxl
import copy

sat_sheet_path = os.path.join("template_sheet", "sat_I_answer_sheet.xlsx")
psat_sheet_path = os.path.join("template_sheet", "psat_answer_sheet.xlsx")
act_sheet_path = os.path.join("template_sheet", "act_answer_sheet.xlsx")

class Generate_Sheet():

    def __init__(self, file_manager):
        self.file_manager = file_manager

    def generate_test(self, student_id, test_info, test_ori):
        cur_dir = os.getcwd()
        pathlib.Path(cur_dir + '/test_sheets').mkdir(parents=True, exist_ok=True)

        now = datetime.datetime.now()
        cur_date = str(now.month) + "-" + str(now.day) + "-" + str(now.year)
        pathlib.Path(cur_dir + '/test_sheets/' + cur_date).mkdir(parents=True, exist_ok=True)

        student = self.file_manager.get_student(student_id)
        source = None
        dest = None

        filename = student["first_name"] + "_" + student["last_name"] + ":" + test_info + ":answer_sheet"
        filename = "".join(filename.split())
        filename = filename.replace(":", "_")

        if "SATI" in test_info:

            source = os.path.join(cur_dir, sat_sheet_path)

        elif "PSAT" in test_info:

            source = os.path.join(cur_dir, psat_sheet_path)

        elif "ACT" in test_info:
            source = os.path.join(cur_dir, act_sheet_path)

        dest = os.path.join(cur_dir, "test_sheets", cur_date, filename + ".xlsx")
        copyfile(source, dest)

        wb = load_workbook(filename=dest)
        ws = wb["Design"]
        ws["AJ3"] = (student["first_name"] + " " + student["last_name"])
        ws["AQ4"] = student["grade"]
        ws["AJ5"] = student_id
        ws["H14"] = test_ori
        ws["h16"] = cur_date
        packet_id = self.file_manager.get_next_packetID()
        ws["H55"] = packet_id[0] + " " + packet_id[1]
        ws["AF55"] = 1
        ws["h116"] = packet_id[0] + " " + packet_id[1]
        ws["AF116"] = 2

        self.file_manager.create_packetID(packet_id[0] + packet_id[1], student_id, cur_date, test_ori)
        self.file_manager.save_file("packet")
        wb.save(filename=dest)

        return dest
